from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.mail import send_mail
from django.db.models import Q, Count, Avg, Sum
from .forms import RegistrationForm, FeedbackForm
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from .models import FeedbackRequest, Feedback, AIUsageLog
from users.models import Profile, ContractSettings, Department, Company
from django.contrib.auth.models import User
from django.http import JsonResponse
from django.db import OperationalError
from django.views.decorators.http import require_POST
import json, traceback
from django.db import models
import logging
from datetime import date, timedelta
from decimal import Decimal
from .services import FeedbackAnalytics
from .ai_service import OpenRouterService
from .forms import DepartmentForm
from django.utils import timezone


logger = logging.getLogger(__name__)

def is_company_active(user):
    if hasattr(user, 'profile') and user.profile.company_link:
        return user.profile.company_link.is_active
    return True

from .models import PageDescription

def index(request):
    if request.user.is_authenticated:
        return redirect('home')
        
    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        employees = request.POST.get('employees', 'Nepasirinkta')
        message = request.POST.get('message')
        
        subject = f"Nauja užklausa iš Orbigrow.lt: {name}"
        body = f"Vardas: {name}\nEl. paštas: {email}\nTelefonas: {phone}\nDarbuotojų skaičius: {employees}\n\nŽinutė:\n{message}"
        
        try:
            send_mail(
                subject,
                body,
                'info@orbigrow.lt',  # From email
                ['info@orbigrow.lt'], # To email
                fail_silently=False,
            )
            messages.success(request, 'Ačiū! Jūsų užklausa gauta, netrukus su jumis susisieksime.')
        except Exception as e:
            logger.error(f"Klaida siunčiant kontaktų formos el. laišką: {str(e)}")
            messages.error(request, 'Apgailestaujame, įvyko klaida siunčiant žinutę. Bandykite vėliau arba rašykite tiesiogiai info@orbigrow.lt.')
            
        return redirect('index')
        
    page_description = PageDescription.load()
    return render(request, 'index.html', {'page_description': page_description})

def apie_mus(request):
    page_description = PageDescription.load()
    return render(request, 'apie_mus.html', {'page_description': page_description})

def security_page(request):
    page_description = PageDescription.load()
    return render(request, 'saugumas.html', {'page_description': page_description})

from django.contrib.auth.decorators import user_passes_test
@login_required
@user_passes_test(lambda u: u.is_superuser)
def superadmin_descriptions(request):
    page_description = PageDescription.load()
    from .forms import PageDescriptionForm
    if request.method == 'POST':
        form = PageDescriptionForm(request.POST, instance=page_description)
        if form.is_valid():
            form.save()
            messages.success(request, 'Aprašymai sėkmingai atnaujinti.')
            return redirect('superadmin_descriptions')
    else:
        form = PageDescriptionForm(instance=page_description)
    
    return render(request, 'superadmin_descriptions.html', {
        'form': form
    })

@login_required
def home(request):
    feedback_requests = FeedbackRequest.objects.filter(requested_to=request.user, status='pending').select_related('requester', 'requester__profile')
    company_name = ''
    try:
        if request.user.profile.company_link:
            company_name = request.user.profile.company_link.name
    except Profile.DoesNotExist:
        pass
    except OperationalError as e:
        logger.error(f"Database operational error fetching company for {request.user.username}: {e}")
    
    # Recent team activity — visos įmonės veikla
    recent_activity = []
    
    try:
        user_company = request.user.profile.company_link
    except (Profile.DoesNotExist, AttributeError):
        user_company = None
    
    if user_company:
        # Neseniai užbaigti atsiliepimai komandoje
        recent_completed = Feedback.objects.filter(
            feedback_request__requester__profile__company_link=user_company,
            feedback_request__status='completed'
        ).select_related(
            'feedback_request__requested_to',
            'feedback_request__requester'
        ).order_by('-feedback_request__created_at')[:10]
        
        for fb in recent_completed:
            writer = fb.feedback_request.requested_to
            about = fb.feedback_request.requester
            is_self = fb.feedback_request.is_self_initiated
            
            if is_self:
                action_text = f'įvertino {about.get_full_name() or about.username}.'
                person = writer
            else:
                action_text = f'pateikė atsiliepimą apie {about.get_full_name() or about.username}.'
                person = writer
            
            recent_activity.append({
                'initials': (person.first_name[:1] + person.last_name[:1]).upper() if person.first_name and person.last_name else '??',
                'name': person.get_full_name() or person.username,
                'action': action_text,
                'date': fb.feedback_request.created_at,
            })
        
        # Neseniai sukurti prašymai komandoje
        recent_requests = FeedbackRequest.objects.filter(
            requester__profile__company_link=user_company,
            is_self_initiated=False
        ).select_related('requester', 'requested_to').order_by('-created_at')[:10]
        
        for fr in recent_requests:
            person = fr.requester
            recent_activity.append({
                'initials': (person.first_name[:1] + person.last_name[:1]).upper() if person.first_name and person.last_name else '??',
                'name': person.get_full_name() or person.username,
                'action': f'paprašė atsiliepimo iš {fr.requested_to.get_full_name() or fr.requested_to.username}.',
                'date': fr.created_at,
            })
    
    # Sort by date descending, take top 5
    recent_activity.sort(key=lambda x: x['date'], reverse=True)
    recent_activity = recent_activity[:5]
    
    # Dynamic metrics calculation
    pending_tasks_count = feedback_requests.count()
    
    # Užpildytos apklausos, gautos vartotojo (surinkta atsakymų)
    completed_surveys_count = Feedback.objects.filter(
        feedback_request__requester=request.user, 
        feedback_request__status='completed'
    ).count()
    
    # Mano išsiųsta kitiems (komandos nariams išsiųstas ryšys)
    sent_surveys_count = Feedback.objects.filter(
        feedback_request__requested_to=request.user,
        feedback_request__status='completed'
    ).count()

    # Calculate distinct years for the feedback chart filter
    current_year = date.today().year
    try:
        selected_year = int(request.GET.get('year', current_year))
    except ValueError:
        selected_year = current_year

    years = FeedbackRequest.objects.filter(requester=request.user).dates('created_at', 'year')
    available_years = sorted(list(set([y.year for y in years] + [current_year])), reverse=True)

    context = {
        'feedback_requests': feedback_requests,
        'company_name': company_name,
        'recent_activity': recent_activity,
        'is_company_active': is_company_active(request.user),
        'pending_tasks_count': pending_tasks_count,
        'completed_surveys_count': completed_surveys_count,
        'sent_surveys_count': sent_surveys_count,
        'available_years': available_years,
        'selected_year': selected_year,
    }
    return render(request, 'home.html', context)

def register(request):
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            company_name = form.cleaned_data.get('company')
            company = None
            if company_name:
                company, created = Company.objects.get_or_create(name=company_name)
            Profile.objects.create(user=user, company_link=company)
            login(request, user, backend='django.contrib.auth.backends.ModelBackend')
            return redirect('home')
    else:
        form = RegistrationForm()
    return render(request, 'registration/register.html', {'form': form})

def logout_view(request):
    logout(request)
    return redirect('index')

@login_required
def get_team_members(request):
    user = request.user
    team_members_qs = User.objects.none()
    try:
        user_company_link = user.profile.company_link
        if user_company_link:
            team_members_qs = User.objects.filter(profile__company_link=user_company_link).exclude(id=user.id)
    except Profile.DoesNotExist:
        pass  # Jei nėra įmonės, tiesiog grąžinsime tuščią sąrašą
    except OperationalError as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Database error fetching team_members for {user.username}: {e}")
        
    from feedbackas.converters import HashIdConverter
    converter = HashIdConverter()
    
    # Gauti narių ID, kuriems jau yra laukianti užklausa
    pending_requester_ids = set(
        FeedbackRequest.objects.filter(
            requested_to=user,
            status='pending'
        ).values_list('requester_id', flat=True)
    )
    
    data = [
        {
            'id': converter.to_url(member.id), 
            'name': member.get_full_name() or member.username,
            'has_pending': member.id in pending_requester_ids
        } 
        for member in team_members_qs
    ]
    return JsonResponse(data, safe=False)

@login_required
def request_feedback(request):
    if not is_company_active(request.user):
        return JsonResponse({'success': False, 'errors': 'Jūsų įmonė yra išjungta. Veiksmas negalimas.'})
        
    if request.method == 'POST':
        requester = request.user
        requested_to_ids = request.POST.getlist('requested_to')
        project_name = request.POST.get('project_name')
        comment = request.POST.get('comment')
        due_date = request.POST.get('due_date')
        
        feedback_request_ids = []
        skipped_names = []
        for requested_to_id in requested_to_ids:
            requested_to = get_object_or_404(User, id=requested_to_id)
            
            if requested_to.profile.company_link != request.user.profile.company_link:
                # Gynyba naršyklės DOM inspekcijoms
                continue
            
            # Patikrinti, ar jau yra neužpildytas prašymas šiam žmogui
            existing_pending = FeedbackRequest.objects.filter(
                requester=requester,
                requested_to=requested_to,
                status='pending'
            ).exists()
            if existing_pending:
                skipped_names.append(requested_to.get_full_name() or requested_to.username)
                continue
            
            feedback_request = FeedbackRequest.objects.create(
                requester=requester,
                requested_to=requested_to,
                project_name=project_name,
                comment=comment,
                due_date=due_date
            )
            feedback_request_ids.append(feedback_request.id)
            
        from feedbackas.converters import HashIdConverter
        converter = HashIdConverter()
        encoded_ids = [converter.to_url(fid) for fid in feedback_request_ids]
        response_data = {'success': True, 'feedback_request_ids': encoded_ids}
        if skipped_names:
            response_data['skipped'] = skipped_names
            response_data['skipped_message'] = f'Praleisti nariai (jau turi laukiančią užklausą): {", ".join(skipped_names)}'
        return JsonResponse(response_data)
    return JsonResponse({'success': False, 'errors': 'Invalid request method'})

@login_required
def send_feedback(request, user_id):
    if not is_company_active(request.user):
        messages.error(request, 'Jūsų įmonė yra išjungta. Veiksmas negalimas.')
        return redirect('home')
        
    requester = get_object_or_404(User, id=user_id)
    requested_to = request.user
    
    # Patikrinti, ar jau yra neužpildytas prašymas nuo šio žmogaus
    existing_pending = FeedbackRequest.objects.filter(
        requester=requester,
        requested_to=requested_to,
        status='pending'
    ).exists()
    if existing_pending:
        messages.warning(request, f'Jūs jau turite neužpildytą atsiliepimo užklausą nuo {requester.get_full_name() or requester.username}. Pirmiau užpildykite esamą.')
        return redirect('home')
    
    feedback_request = FeedbackRequest.objects.create(
        requester=requester,
        requested_to=requested_to,
        project_name='Atsiliepimas',
        comment='',
        due_date=date.today(),
        is_self_initiated=True
    )
    
    return redirect('fill_feedback', request_id=feedback_request.id)

@login_required
def fill_feedback(request, request_id):
    if not is_company_active(request.user):
        messages.error(request, 'Jūsų įmonė yra išjungta. Veiksmas negalimas.')
        return redirect('home')
        
    feedback_request = get_object_or_404(FeedbackRequest, id=request_id)
    
    if feedback_request.requested_to != request.user:
        messages.error(request, 'Neturite teisių pildyti šio atsiliepimo.')
        return redirect('home')
    if request.method == 'POST':
        form = FeedbackForm(request.POST)
        if form.is_valid():
            feedback = form.save(commit=False)
            feedback.feedback_request = feedback_request
            
            # Užtikriname, kad pakoreguotas tekstas būtų paimtas, jei form.save(commit=False) jo neužpildė
            if 'feedback' in request.POST:
                feedback.feedback = request.POST.get('feedback')
            
            feedback.save()
            
            # Save trait ratings if this is a questionnaire-based feedback
            if feedback_request.questionnaire:
                from .models import TraitRating
                for trait in feedback_request.questionnaire.traits.all():
                    trait_rating_value = request.POST.get(f'trait_rating_{trait.id}', 0)
                    try:
                        trait_rating_value = int(trait_rating_value)
                    except (ValueError, TypeError):
                        trait_rating_value = 0
                    TraitRating.objects.update_or_create(
                        feedback=feedback,
                        trait=trait,
                        defaults={'rating': trait_rating_value}
                    )
            
            feedback_request.status = 'completed'
            feedback_request.save()
            
            # AI Išskyrimas (Stiprybės ir Tobulintinos sritys) - Foninė užduotis
            # Perkeliame čia, kad užtikrintume, jog feedback.feedback jau yra DB
            from django_q.tasks import async_task
            async_task('feedbackas.services.extract_feedback_features_task', feedback.id)
            
            messages.success(request, 'Jūsų įvertinimas išsiųstas.')
            return redirect('home')
    else:
        form = FeedbackForm()
    
    # If this feedback request is linked to a questionnaire, pass its traits
    import json
    questionnaire_traits = []
    if feedback_request.questionnaire:
        questionnaire_traits = [
            {'id': t.id, 'name': t.name}
            for t in feedback_request.questionnaire.traits.all()
        ]
    
    is_team_form = False
    if feedback_request.questionnaire and ' (' in feedback_request.project_name and feedback_request.project_name.endswith(')'):
        is_team_form = True

    context = {
        'form': form,
        'feedback_request': feedback_request,
        'questionnaire_traits_json': json.dumps(questionnaire_traits),
        'has_questionnaire': feedback_request.questionnaire is not None,
        'is_team_form': is_team_form,
    }
    return render(request, 'fill_feedback.html', context)



@login_required
def team_members_list(request):
    user = request.user
    team_members_qs = User.objects.none() 

    try:
        user_department = user.profile.department
        user_company_link = user.profile.company_link
        
        # Check if user manages any department with sub-departments
        managed_departments = Department.objects.filter(manager=user)
        sub_departments = Department.objects.filter(parent__in=managed_departments)
        
        if sub_departments.exists():
            # Hierarchical mode: show department blocks
            department_blocks = []
            
            # Sub-departments loop optimization
            for dept in sub_departments.select_related('manager').prefetch_related('members__user'):
                members_qs = User.objects.filter(profile__department=dept).select_related('profile').annotate(
                    average_rating=Avg('made_requests__feedback__rating')
                )
                
                # Fetch department avg with a single database hit using filter instead of evaluating qs
                dept_avg = Feedback.objects.filter(
                    feedback_request__requester__profile__department=dept,
                    feedback_request__status='completed'
                ).aggregate(Avg('rating'))['rating__avg']
                
                department_blocks.append({
                    'department': dept,
                    'members': members_qs,
                    'member_count': members_qs.count(),
                    'avg_rating': round(dept_avg, 2) if dept_avg else None,
                })
            
            # Also include direct members of the managed department (not in sub-depts)
            for managed_dept in managed_departments:
                direct_members = User.objects.filter(profile__department=managed_dept).exclude(id=user.id).select_related('profile').annotate(
                    average_rating=Avg('made_requests__feedback__rating')
                )
                if direct_members.exists():
                    direct_avg = Feedback.objects.filter(
                        feedback_request__requester__profile__department=managed_dept,
                        feedback_request__status='completed'
                    ).exclude(feedback_request__requester=user).aggregate(Avg('rating'))['rating__avg']
                    
                    department_blocks.insert(0, {
                        'department': managed_dept,
                        'members': direct_members,
                        'member_count': direct_members.count(),
                        'avg_rating': round(direct_avg, 2) if direct_avg else None,
                    })
            
            # Overall stats
            all_members = User.objects.filter(
                Q(profile__department__in=sub_departments) | 
                Q(profile__department__in=managed_departments)
            ).exclude(id=user.id)
            overall_avg_rating = Feedback.objects.filter(
                feedback_request__requester__in=all_members, 
                feedback_request__status='completed'
            ).aggregate(Avg('rating'))['rating__avg']
            pending_feedback_count = FeedbackRequest.objects.filter(
                requester__in=all_members, status='pending'
            ).count()
            
            # ID sąrašas narių, kuriems jau išsiųsta laukianti užklausa
            pending_from_me_ids = set(
                FeedbackRequest.objects.filter(
                    requested_to=user,
                    status='pending'
                ).values_list('requester_id', flat=True)
            )
            
            context = {
                'has_sub_departments': True,
                'department_blocks': department_blocks,
                'team_members': all_members,
                'search_query': None,
                'overall_avg_rating': overall_avg_rating,
                'pending_feedback_count': pending_feedback_count,
                'pending_from_me_ids': pending_from_me_ids,
            }
            return render(request, 'feedbackas/team_members_list.html', context)
        
        # Flat mode: show single department members
        if user_department:
            team_members_qs = User.objects.filter(profile__department=user_department).exclude(id=user.id)
        elif user_company_link:
            team_members_qs = User.objects.filter(profile__company_link=user_company_link, profile__department__isnull=True).exclude(id=user.id)
        else:
            team_members_qs = User.objects.filter(profile__company_link__isnull=True).exclude(id=user.id)
    except Profile.DoesNotExist:
        team_members_qs = User.objects.filter(profile__isnull=True).exclude(id=user.id)

    # Paieškos logika
    query = request.GET.get('q')
    if query:
        team_members_qs = team_members_qs.filter(
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query)
        )

    # Anotuojame su papildomais duomenimis
    team_members = team_members_qs.select_related('profile').annotate(
        average_rating=Avg('made_requests__feedback__rating')
    )

    # Bendra statistika
    overall_avg_rating = Feedback.objects.filter(feedback_request__requester__in=team_members_qs).aggregate(Avg('rating'))['rating__avg']
    pending_feedback_count = FeedbackRequest.objects.filter(requester__in=team_members_qs, status='pending').count()

    # ID sąrašas narių, kuriems jau išsiųsta laukianti užklausa
    pending_from_me_ids = set(
        FeedbackRequest.objects.filter(
            requested_to=user,
            status='pending'
        ).values_list('requester_id', flat=True)
    )
    
    context = {
        'has_sub_departments': False,
        'team_members': team_members,
        'search_query': query,
        'overall_avg_rating': overall_avg_rating,
        'pending_feedback_count': pending_feedback_count,
        'pending_from_me_ids': pending_from_me_ids,
    }
    
    return render(request, 'feedbackas/team_members_list.html', context)

@login_required
def my_tasks_list(request):
    # Feedback requests made by the current user (excluding self-initiated evaluations from others)
    made_requests = FeedbackRequest.objects.filter(requester=request.user, is_self_initiated=False).select_related('requested_to', 'feedback').order_by('-due_date')

    # Feedback requests assigned to the current user (tasks to do)
    assigned_requests = FeedbackRequest.objects.filter(requested_to=request.user).select_related('requester').order_by('-due_date')

    context = {
        'made_requests': made_requests,
        'assigned_requests': assigned_requests,
    }
    return render(request, 'my_tasks.html', context)

@login_required
@require_POST
def cancel_feedback_request(request, request_id):
    """
    Suteikia galimybę ištrinti prašymą, kurį sukūrė vartotojas.
    """
    feedback_request = get_object_or_404(FeedbackRequest, id=request_id)

    # Patikriname ar esamas vartotojas yra prašymo autorius
    if feedback_request.requester != request.user:
        messages.error(request, 'Neturite teisių ištrinti šio prašymo.')
        return redirect('my_tasks_list')
        
    feedback_request.delete()
    messages.success(request, 'Atsiliepimo prašymas sėkmingai ištrintas.')
    return redirect('my_tasks_list')

@login_required
@require_POST
def reject_feedback_request(request, request_id):
    """
    Leidžia vartotojui (kurio prašoma atsiliepimo) ištrinti (atmesti) prašymą.
    """
    feedback_request = get_object_or_404(FeedbackRequest, id=request_id)

    if feedback_request.requested_to != request.user:
        messages.error(request, 'Neturite teisių ištrinti šio prašymo.')
        return redirect('my_tasks_list')
        
    feedback_request.delete()
    messages.success(request, 'Atsiliepimo užklausa sėkmingai ištrinta.')
    
    return redirect('my_tasks_list')

@login_required
def edit_feedback_request(request, request_id):
    """
    Leidžia vartotojui paredaguoti prašymo projekto pavadinimą, komentarą ir terminą.
    """
    feedback_request = get_object_or_404(FeedbackRequest, id=request_id)

    # Autoriaus ir statuso patikrinimai
    if feedback_request.requester != request.user:
        messages.error(request, 'Neturite teisių redaguoti šio prašymo.')
        return redirect('my_tasks_list')
        
    if feedback_request.status != 'pending':
        messages.error(request, 'Begalima redaguoti jau įvertinto arba užbaigto prašymo.')
        return redirect('my_tasks_list')
        
    if request.method == 'POST':
        project_name = request.POST.get('project_name')
        comment = request.POST.get('comment')
        due_date = request.POST.get('due_date')
        
        if project_name and due_date:
            feedback_request.project_name = project_name
            feedback_request.comment = comment
            feedback_request.due_date = due_date
            feedback_request.save()
            messages.success(request, 'Atsiliepimo prašymas sėkmingai atnaujintas.')
        else:
            messages.error(request, 'Užpildykite visus privalomus laukus (Projekto pavadinimas, Terminas).')

    return redirect('my_tasks_list')




from django_ratelimit.decorators import ratelimit

from django_q.tasks import async_task, result

@login_required
@require_POST
@ratelimit(key='user', rate='10/10m', block=True)
def generate_ai_feedback(request):
    try:
        data = json.loads(request.body)
        ratings = data.get('ratings', {})
        keywords = data.get('keywords', '')
        comments = data.get('comments', '')
        existing_feedback = data.get('existing_feedback', '')
        colleague_name = data.get('colleague_name', 'Kolega')

        task_id = async_task(
            'feedbackas.services.generate_ai_feedback_task',
            ratings=ratings,
            keywords=keywords,
            comments=comments,
            existing_feedback=existing_feedback,
            colleague_name=colleague_name,
            user_id=request.user.id
        )
        
        return JsonResponse({'task_id': task_id, 'status': 'processing'})

    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        logger.error(f"AI feedback dispatch failed: {e}\n{traceback.format_exc()}")
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def check_ai_task_status(request):
    task_id = request.GET.get('task_id')
    if not task_id:
        return JsonResponse({'error': 'No task_id provided'}, status=400)
        
    try:
        from django_q.models import Task
        task = Task.get_task(task_id)
        if task is None:
            return JsonResponse({'status': 'processing'})
            
        if task.success:
            return JsonResponse({'status': 'completed', 'generated_feedback': task.result})
        else:
            return JsonResponse({'status': 'failed', 'error': 'Task failed to execute'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def get_feedback_data(request):
    user = request.user
    
    current_year = date.today().year
    try:
        year = int(request.GET.get('year', current_year))
    except ValueError:
        year = current_year
        
    # Get all feedback requests made by this user in the specified year, ordered by creation date
    all_requests = FeedbackRequest.objects.filter(
        requester=user,
        created_at__year=year
    ).select_related('requested_to', 'feedback').order_by('created_at')
    
    from feedbackas.converters import HashIdConverter
    converter = HashIdConverter()
    
    data = []
    for fr in all_requests:
        respondent = fr.requested_to.get_full_name() or fr.requested_to.username
        label = f"{fr.project_name} ({respondent})"
        
        feedback_details = None
        if fr.status == 'completed':
            status = 'done'
            try:
                # Accessing a missing OneToOneField can raise RelatedObjectDoesNotExist
                if fr.feedback:
                    feedback_details = {
                        'project': fr.project_name,
                        'rname': respondent,
                        'rating': fr.feedback.rating,
                        'date': fr.feedback.created_at.strftime('%Y-%m-%d'),
                        'feedback': fr.feedback.feedback,
                        'comments': fr.feedback.comments,
                    }
            except Exception:
                pass
        elif fr.status == 'pending':
            status = 'active'
        else:
            status = 'empty'
            
        data.append({
            'id': converter.to_url(fr.id),
            'label': label,
            'status': status,
            'feedback_details': feedback_details
        })
    
    # Tikslas: 10 apklausų, todėl turi matytis 10 tuščių (arba pilnų) burbuliukų
    while len(data) < 10:
        data.append({
            'id': f'empty-{len(data)}',
            'label': 'Apklausa',
            'status': 'upcoming',
            'feedback_details': None
        })
        
    return JsonResponse(data, safe=False)


@login_required
def results(request):
    user = request.user
    period = request.GET.get('period', 'all')
    stats = FeedbackAnalytics.get_user_stats(user, period=period)
    
    company_name = ''
    if hasattr(request.user, 'profile') and request.user.profile.company_link:
        company_name = request.user.profile.company_link.name

    context = {
        **stats,
        'company_name': company_name,
        'current_period': period,
    }
    
    return render(request, 'results.html', context)


@login_required
def get_competency_trend(request, competency_name):
    """
    Returns the historical evaluation scores for a specific competency for the current user.
    Uses the actual DB rating fields (teamwork_rating, communication_rating, etc.)
    """
    user = request.user

    # Map Lithuanian competency display names to DB field names
    competency_field_map = {
        'komandinis darbas': 'teamwork_rating',
        'komunikacija': 'communication_rating',
        'iniciatyvumas': 'initiative_rating',
        'techninės žinios': 'technical_skills_rating',
        'problemų sprendimas': 'problem_solving_rating',
    }

    field_name = competency_field_map.get(competency_name.strip().lower())
    if not field_name:
        return JsonResponse({'competency': competency_name, 'trend': []})

    # Get all completed feedback for this user (they are the requester)
    feedbacks = Feedback.objects.filter(
        feedback_request__requester=user,
        feedback_request__status='completed'
    ).select_related('feedback_request').order_by('feedback_request__created_at')

    trend_data = []
    for fb in feedbacks:
        score_val = getattr(fb, field_name, None)
        if score_val is not None:
            trend_data.append({
                'date': fb.feedback_request.created_at.strftime('%Y-%m-%d'),
                'score': float(score_val),
                'project': fb.feedback_request.project_name
            })

    return JsonResponse({'competency': competency_name, 'trend': trend_data})


@login_required
def team_statistics(request):
    user = request.user
    
    # Find departments managed by this user
    managed_departments = Department.objects.filter(manager=user)
    if not managed_departments.exists():
        from django.contrib import messages as django_messages
        django_messages.warning(request, 'Jūs nesate jokio padalinio vadovas.')
        return redirect('home')
    
    department = managed_departments.first()
    team_members = User.objects.filter(profile__department=department).exclude(id=user.id)
    
    # Aggregate stats using TeamAnalytics service
    from .services import TeamAnalytics
    stats = TeamAnalytics.get_team_stats(team_members)

    context = {
        'department': department,
        'member_stats': stats['member_stats'],
        'team_avg_rating': round(stats['team_avg_rating'], 2),
        'team_feedback_count': stats['team_feedback_count'],
        'team_member_count': stats['team_member_count'],
        'competencies': stats['competencies'],
    }
    return render(request, 'team_statistics.html', context)

@login_required
def team_member_detail(request, user_id):
    member = get_object_or_404(User, id=user_id)
    
    # Verify current user is a manager of the member's department or a parent department
    member_dept = member.profile.department if hasattr(member, 'profile') else None
    is_authorized = False
    if member_dept:
        # Check direct manager
        if member_dept.manager == request.user:
            is_authorized = True
        else:
            # Walk up the parent chain
            parent = member_dept.parent
            while parent:
                if parent.manager == request.user:
                    is_authorized = True
                    break
                parent = parent.parent
    if not is_authorized:
        from django.contrib import messages as django_messages
        django_messages.error(request, 'Jūs neturite teisės peržiūrėti šio darbuotojo informacijos.')
        return redirect('home')
    
    # All completed feedback about this member
    feedbacks = Feedback.objects.filter(
        feedback_request__requester=member,
        feedback_request__status='completed'
    ).select_related('feedback_request', 'feedback_request__requested_to').order_by('-feedback_request__created_at')
    
    # Aggregate stats using TeamAnalytics service
    from .services import TeamAnalytics
    stats = TeamAnalytics.get_member_detailed_stats(feedbacks)
    
    # Trait ratings (from questionnaire-based feedback)
    from .models import TraitRating
    trait_ratings = TraitRating.objects.filter(
        feedback__feedback_request__requester=member
    ).select_related('trait').values('trait__name').annotate(
        avg_rating=Avg('rating')
    ).order_by('-avg_rating')
    
    context = {
        'member': member,
        'department': member_dept,
        'feedbacks': feedbacks,
        'avg_rating': stats['avg_rating'],
        'feedback_count': feedbacks.count(),
        'competencies': stats['competencies'],
        'all_keywords': list(stats['keywords'])[:15],
        'trait_ratings': trait_ratings,
    }
    return render(request, 'team_member_detail.html', context)

@login_required
def all_feedback_list(request):
    # Fetch only feedback received about the current user
    all_feedback = Feedback.objects.select_related(
        'feedback_request__requester', 
        'feedback_request__requested_to'
    ).filter(
        feedback_request__requester=request.user,
        feedback_request__status='completed'
    ).order_by('-feedback_request__created_at')

    context = {
        'all_feedback': all_feedback,
    }
    return render(request, 'all_feedback_list.html', context)

@login_required
def company_management(request):
    user = request.user
    
    # Saugumas: Patikrinimas ar vartotojas turi teises
    if not user.is_superuser and not getattr(user.profile, 'is_company_admin', False):
        from django.contrib import messages
        messages.error(request, 'Neturite teisių valdyti įmonės struktūros.')
        return redirect('home')

    try:
        user_company = user.profile.company_link
    except AttributeError:
        # Fallback jei dar nėra susieto Company objekto
        return render(request, 'company_management.html', {'error': 'Jūs nepriskirtas jokiai įmonei.'})

    if not user_company:
         return redirect('home') # Arba error page

    # Formos apdorojimas (Pridėti departamentą)
    if request.method == 'POST':
        form = DepartmentForm(user, request.POST)
        if form.is_valid():
            department = form.save(commit=False)
            department.company = user_company
            department.save()
            # Auto-assign manager to this department
            if department.manager:
                manager_profile = department.manager.profile
                manager_profile.department = department
                manager_profile.save()
            return redirect('company_management')
    else:
        form = DepartmentForm(user)

    # Gauname tik šaknijinius departamentus (kurie neturi tėvo)
    # Vaikus gausime template su rekursija arba prefetch_related
    root_departments = Department.objects.filter(company=user_company, parent__isnull=True).prefetch_related('sub_departments')
    
    # Gauname darbuotojus be departamento, kad galėtume juos priskirti
    unassigned_users = User.objects.filter(profile__company_link=user_company, profile__department__isnull=True)

    # All departments for the assignment dropdown
    all_departments = Department.objects.filter(company=user_company).order_by('name')

    context = {
        'root_departments': root_departments,
        'form': form,
        'unassigned_users': unassigned_users,
        'company_name': user_company.name,
        'all_departments': all_departments,
    }
    return render(request, 'company_management.html', context)

@login_required
def assign_to_department(request):
    if request.method == 'POST':
        user = request.user
        
        # Saugumas: Patikrinimas ar vartotojas turi teises
        if not user.is_superuser and not getattr(user.profile, 'is_company_admin', False):
            from django.contrib import messages
            messages.error(request, 'Neturite teisių atlikti šio veiksmo.')
            return redirect('home')
            
        user_id = request.POST.get('user_id')
        department_id = request.POST.get('department_id')
        if user_id and department_id:
            target_user = get_object_or_404(User, id=user_id)
            department = get_object_or_404(Department, id=department_id)
            # Ensure both belong to the same company as the current user
            user_company = request.user.profile.company_link
            if department.company == user_company and target_user.profile.company_link == user_company:
                target_user.profile.department = department
                target_user.profile.save()
    return redirect('company_management')

from django.contrib.auth.decorators import user_passes_test

@user_passes_test(lambda u: u.is_superuser)
def superadmin_dashboard(request):
    # Statistics
    total_users = User.objects.count()
    total_companies = Company.objects.count()
    pending_feedback_count = FeedbackRequest.objects.filter(status='pending').count()
    completed_feedback_count = FeedbackRequest.objects.filter(status='completed').count()

    now = timezone.now()
    
    # 1. AI Costs for current month
    total_ai_cost = AIUsageLog.objects.filter(
        timestamp__year=now.year, 
        timestamp__month=now.month
    ).aggregate(Sum('total_cost'))['total_cost__sum'] or 0.0

    # 2. Revenue for current month
    # This is an approximation: sum(company_employee_count * price_per_employee)
    total_monthly_revenue = Decimal('0.00')
    active_contracts = ContractSettings.objects.filter(
        models.Q(contract_end__isnull=True) | models.Q(contract_end__gte=now.date()),
        contract_start__lte=now.date()
    ).select_related('company')

    for contract in active_contracts:
        employee_count = Profile.objects.filter(company_link=contract.company).count()
        revenue = Decimal(str(employee_count)) * contract.price_per_employee
        # Ensure it's at least the minimum fee
        total_monthly_revenue += max(revenue, contract.minimum_fee)

    # Calculate Yearly Totals
    start_of_year = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    
    yearly_total_ai_cost = AIUsageLog.objects.filter(
        timestamp__year=now.year
    ).aggregate(Sum('total_cost'))['total_cost__sum'] or 0.0

    yearly_new_users = User.objects.filter(date_joined__gte=start_of_year).count()
    yearly_new_companies = Company.objects.filter(created_at__gte=start_of_year).count()
    
    yearly_completed_feedback = Feedback.objects.filter(
        created_at__gte=start_of_year
    ).count()

    # Yearly revenue estimation (simplistic: current monthly revenue * months elapsed)
    # Better: sum historical records if available, but here we'll just show current yearly progress
    yearly_total_revenue = total_monthly_revenue * now.month

    # Generate monthly history for charts (last 6 months)
    months_labels = []
    revenue_history = []
    ai_cost_history = []
    users_history = []
    companies_history = []
    feedback_history = []

    for i in range(5, -1, -1):
        target_date = now - timedelta(days=i*30)
        m = target_date.month
        y = target_date.year
        months_labels.append(target_date.strftime('%b'))

        # Revenue (simplified for history: uses current monthly logic for each month)
        # Note: In a real app, you'd query historical snapshots or invoice totals
        # Here we'll just simulate a slight trend for visual effect
        revenue_history.append(float(total_monthly_revenue) * (1 - (i * 0.05))) 

        # AI Cost
        ai_m = AIUsageLog.objects.filter(timestamp__year=y, timestamp__month=m).aggregate(Sum('total_cost'))['total_cost__sum'] or 0.0
        ai_cost_history.append(float(ai_m))

        # Users
        u_m = User.objects.filter(date_joined__year=y, date_joined__month=m).count()
        users_history.append(u_m)

        # Companies
        c_m = Company.objects.filter(created_at__year=y, created_at__month=m).count()
        companies_history.append(c_m)

        # Completed Feedback
        f_m = Feedback.objects.filter(created_at__year=y, created_at__month=m).count()
        feedback_history.append(f_m)

    # Recent Data
    recent_users = User.objects.order_by('-date_joined')[:5]
    recent_companies = Company.objects.order_by('-created_at')[:5]

    context = {
        'total_users': total_users,
        'total_companies': total_companies,
        'pending_feedback_count': pending_feedback_count,
        'completed_feedback_count': completed_feedback_count,
        'recent_users': recent_users,
        'recent_companies': recent_companies,
        'total_monthly_revenue': total_monthly_revenue,
        'total_ai_cost': total_ai_cost,
        'yearly_total_revenue': yearly_total_revenue,
        'yearly_total_ai_cost': yearly_total_ai_cost,
        'yearly_new_users': yearly_new_users,
        'yearly_new_companies': yearly_new_companies,
        'yearly_completed_feedback': yearly_completed_feedback,
        'months_labels': months_labels,
        'revenue_history': revenue_history,
        'ai_cost_history': ai_cost_history,
        'users_history': users_history,
        'companies_history': companies_history,
        'feedback_history': feedback_history,
    }
    return render(request, 'superadmin/dashboard.html', context)

@user_passes_test(lambda u: u.is_superuser)
def superadmin_companies_list(request):
    companies = Company.objects.annotate(employee_count=Count('profile')).order_by('-created_at')
    
    context = {
        'companies': companies,
    }
    return render(request, 'superadmin/companies_list.html', context)

@user_passes_test(lambda u: u.is_superuser)
def superadmin_ai_analytics(request):
    from datetime import datetime, timedelta
    from django.db.models import Sum, Count
    from feedbackas.models import AIUsageLog

    # Determine date range
    today = timezone.now().date()
    # Default: this month
    first_day_of_month = today.replace(day=1)
    
    start_date_str = request.GET.get('start_date', first_day_of_month.strftime('%Y-%m-%d'))
    end_date_str = request.GET.get('end_date', today.strftime('%Y-%m-%d'))

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        start_date = first_day_of_month
        end_date = today

    # Filter logs
    # Make sure we include the end date completely
    end_date_inclusive = end_date + timedelta(days=1)
    logs = AIUsageLog.objects.filter(timestamp__gte=start_date, timestamp__lt=end_date_inclusive)

    # Global KPI
    total_cost = logs.aggregate(Sum('total_cost'))['total_cost__sum'] or 0.0
    total_queries = logs.count()

    # Aggregate by company
    company_stats = logs.values('company__name').annotate(
        total_cost=Sum('total_cost'),
        total_queries=Count('id')
    ).order_by('-total_cost')

    company_labels = []
    company_costs = []
    for stat in company_stats:
        company_labels.append(stat['company__name'] or 'Nepriskirta įmonė')
        company_costs.append(float(stat['total_cost']))

    # Aggregate by user (Top 20) — neįtraukiame foninių užklausų (feedback_analysis)
    user_stats = logs.exclude(request_type='feedback_analysis').values('user__first_name', 'user__last_name', 'user__username', 'company__name').annotate(
        total_cost=Sum('total_cost'),
        total_queries=Count('id')
    ).order_by('-total_cost')[:20]

    context = {
        'start_date': start_date_str,
        'end_date': end_date_str,
        'total_cost': total_cost,
        'total_queries': total_queries,
        'company_labels': json.dumps(company_labels),
        'company_costs': json.dumps(company_costs),
        'user_stats': user_stats,
    }
    return render(request, 'superadmin/ai_analytics.html', context)


@user_passes_test(lambda u: u.is_superuser)
def superadmin_download_employee_template(request):
    import csv
    from django.http import HttpResponse
    
    response = HttpResponse(
        content_type='text/csv',
        headers={'Content-Disposition': 'attachment; filename="darbuotojai_pavyzdys.csv"'},
    )
    # Add BOM for Excel compatibility with UTF-8
    response.write('\ufeff'.encode('utf8'))
    
    writer = csv.writer(response, delimiter=';')
    writer.writerow(['Vardas', 'Pavardė', 'El. paštas', 'Slaptažodis', 'Komandos pavadinimas'])
    writer.writerow(['Jonas', 'Jonaitis', 'jonas.jonaitis@imone.lt', 'slaptazodis123', 'IT Skyrius'])
    writer.writerow(['Petras', 'Petraitis', 'petras.petraitis@imone.lt', 'slaptazodis456', 'Pardavimai'])
    
    return response

@user_passes_test(lambda u: u.is_superuser)
def superadmin_create_company(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        if name:
            from users.models import Company, Department, Profile
            from django.contrib.auth.models import User
            from django.db import transaction
            import csv
            import io

            try:
                with transaction.atomic():
                    company = Company.objects.create(name=name)
                    employee_file = request.FILES.get('employee_list')
                    
                    if employee_file:
                        file_ext = employee_file.name.split('.')[-1].lower()
                        if file_ext == 'csv':
                            file_data = employee_file.read().decode('utf-8-sig')
                            sniffer = csv.Sniffer()
                            try:
                                dialect = sniffer.sniff(file_data[:1024])
                                csv_data = csv.reader(io.StringIO(file_data), dialect)
                            except csv.Error:
                                if ';' in file_data[:1024]:
                                    csv_data = csv.reader(io.StringIO(file_data), delimiter=';')
                                else:
                                    csv_data = csv.reader(io.StringIO(file_data), delimiter=',')
                                    
                            next(csv_data, None) # Skip header
                            for row in csv_data:
                                if len(row) >= 5:
                                    first_name = row[0].strip()
                                    last_name = row[1].strip()
                                    email = row[2].strip()
                                    password = row[3].strip()
                                    team_name = row[4].strip()
                                    
                                    if not email: continue
                                    
                                    user, created = User.objects.get_or_create(email=email, defaults={
                                        'username': email,
                                        'first_name': first_name,
                                        'last_name': last_name,
                                    })
                                    if created or not user.has_usable_password():
                                        user.set_password(password)
                                        user.save()
                                        
                                    department = None
                                    if team_name:
                                        department, _ = Department.objects.get_or_create(name=team_name, company=company)
                                        
                                    profile, _ = Profile.objects.get_or_create(user=user)
                                    profile.company_link = company
                                    if department:
                                        profile.department = department
                                    profile.save()
                        elif file_ext in ['xlsx', 'xls']:
                            try:
                                import openpyxl
                                wb = openpyxl.load_workbook(employee_file)
                                sheet = wb.active
                                for row in sheet.iter_rows(min_row=2, values_only=True):
                                    if row and len(row) >= 5 and row[2]:
                                        first_name = str(row[0]).strip() if row[0] else ''
                                        last_name = str(row[1]).strip() if row[1] else ''
                                        email = str(row[2]).strip()
                                        password = str(row[3]).strip() if row[3] else ''
                                        team_name = str(row[4]).strip() if row[4] else ''
                                        
                                        user, created = User.objects.get_or_create(email=email, defaults={
                                            'username': email,
                                            'first_name': first_name,
                                            'last_name': last_name,
                                        })
                                        if created or not user.has_usable_password():
                                            user.set_password(password)
                                            user.save()
                                            
                                        department = None
                                        if team_name:
                                            department, _ = Department.objects.get_or_create(name=team_name, company=company)
                                            
                                        profile, _ = Profile.objects.get_or_create(user=user)
                                        profile.company_link = company
                                        if department:
                                            profile.department = department
                                        profile.save()
                            except ImportError:
                                messages.warning(request, f'Įmonė "{name}" sukurta, bet nepavyko apdoroti Excel failo. Instaliuokite "openpyxl".')
                                return redirect('superadmin_companies_list')
                                
                    messages.success(request, f'Įmonė "{name}" sėkmingai sukurta.')
            except Exception as e:
                messages.error(request, f'Įvyko klaida: {str(e)}')
            return redirect('superadmin_companies_list')
    return render(request, 'superadmin/company_create.html')

@user_passes_test(lambda u: u.is_superuser)
def superadmin_company_detail(request, company_id):
    company = get_object_or_404(Company, id=company_id)
    employee_count = company.profile_set.count()
    department_count = company.departments.count()
    employees = Profile.objects.filter(company_link=company).select_related('user', 'department').order_by('-is_company_admin', 'user__first_name')

    context = {
        'company': company,
        'employee_count': employee_count,
        'department_count': department_count,
        'employees': employees,
    }
    return render(request, 'superadmin/company_detail.html', context)

@user_passes_test(lambda u: u.is_superuser)
@require_POST
def superadmin_toggle_company(request, company_id):
    company = get_object_or_404(Company, id=company_id)
    company.is_active = not company.is_active
    company.save()
    status = 'įjungta' if company.is_active else 'išjungta'
    messages.success(request, f'Įmonė "{company.name}" apskaitos būsena pakeista į: {status}.')
    return redirect('superadmin_companies_list')

@user_passes_test(lambda u: u.is_superuser)
@require_POST
def superadmin_delete_company(request, company_id):
    company = get_object_or_404(Company, id=company_id)
    name = company.name
    # Prieš trinant įmonę patikriname, ar reikia išvalyti vartotojus:
    # Company model is linked to Profile. CASCADE delete will delete profiles if models.CASCADE is set.
    # But usually it's models.SET_NULL or CASCADE based on users/models.py logic.
    # In Profile it is models.SET_NULL. Thus users are kept but unlinked.
    # For now we'll just delete the Company object safely.
    company.delete()
    messages.success(request, f'Įmonė "{name}" sėkmingai ištrinta.')
    return redirect('superadmin_companies_list')

@user_passes_test(lambda u: u.is_superuser)
def superadmin_edit_hierarchy(request, company_id):
    company = get_object_or_404(Company, id=company_id)
    
    # Passing a dummy user to DepartmentForm init is necessary because it expects a user 
    # to filter queryset for parent departments. 
    # Ideally DepartmentForm should be refactored to accept querysets directly, 
    # but for now we can rely on how it filters using user.profile.company_link.
    # HOWEVER, since we are superadmin, we don't have a company link matching the target company necessarily.
    # We need to instantiate the form and then manually override the parent queryset.
    
    if request.method == 'POST':
        form = DepartmentForm(request.user, request.POST)
        # Override parent and manager querysets to target company
        form.fields['parent'].queryset = Department.objects.filter(company=company)
        form.fields['manager'].queryset = User.objects.filter(profile__company_link=company)
        
        if form.is_valid():
            department = form.save(commit=False)
            department.company = company
            department.save()
            # Auto-assign manager to this department
            if department.manager:
                manager_profile = department.manager.profile
                manager_profile.department = department
                manager_profile.save()
            return redirect('superadmin_edit_hierarchy', company_id=company_id)
    else:
        form = DepartmentForm(request.user)
        form.fields['parent'].queryset = Department.objects.filter(company=company)
        form.fields['manager'].queryset = User.objects.filter(profile__company_link=company)

    root_departments = Department.objects.filter(company=company, parent__isnull=True).prefetch_related('sub_departments')

    context = {
        'company': company,
        'form': form,
        'root_departments': root_departments,
    }
    return render(request, 'superadmin/edit_hierarchy.html', context)

@user_passes_test(lambda u: u.is_superuser)
def superadmin_edit_employees(request, company_id):
    company = get_object_or_404(Company, id=company_id)
    profiles = Profile.objects.filter(company_link=company).select_related('user', 'department', 'manager')
    departments = Department.objects.filter(company=company)
    
    context = {
        'company': company,
        'profiles': profiles,
        'departments': departments,
    }
    return render(request, 'superadmin/edit_employees.html', context)

@user_passes_test(lambda u: u.is_superuser)
def superadmin_add_employee(request, company_id):
    if request.method == 'POST':
        email = request.POST.get('email')
        first_name = request.POST.get('first_name', '')
        last_name = request.POST.get('last_name', '')
        password = request.POST.get('password')
        department_id = request.POST.get('department_id')
        
        company = get_object_or_404(Company, id=company_id)
        department = Department.objects.filter(id=department_id, company=company).first() if department_id else None

        try:
            user = User.objects.get(email=email)
            
            # Update user details if provided
            updated = False
            if first_name:
                user.first_name = first_name
                updated = True
            if last_name:
                user.last_name = last_name
                updated = True
            if password:
                user.set_password(password)
                updated = True
            if updated:
                user.save()

            if hasattr(user, 'profile'):
                if user.profile.company_link and user.profile.company_link != company:
                    messages.error(request, f'Vartotojas {email} jau priklauso kitai įmonei.')
                else:
                    user.profile.company_link = company
                    user.profile.department = department
                    user.profile.save()
                    messages.success(request, f'Vartotojas {email} sėkmingai atnaujintas / pridėtas prie įmonės.')
            else:
                 # If user has no profile, create one
                Profile.objects.create(user=user, company_link=company, department=department)
                messages.success(request, f'Vartotojas {email} sėkmingai pridėtas prie įmonės.')

        except User.DoesNotExist:
            # Create a new user
            username = email.split('@')[0]
            # Handle potential username conflicts
            if User.objects.filter(username=username).exists():
                import uuid
                username = f"{username}_{str(uuid.uuid4())[:8]}"
            
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password if password else User.objects.make_random_password(),
                first_name=first_name,
                last_name=last_name
            )
            
            # The Profile is usually auto-created by signals, so we retrieve or create it to avoid IntegrityError
            profile, _ = Profile.objects.get_or_create(user=user)
            profile.company_link = company
            profile.department = department
            profile.save()
            
            messages.success(request, f'Naujas vartotojas {email} sėkmingai sukurtas ir pridėtas.')
    
    return redirect('superadmin_edit_employees', company_id=company_id)

@user_passes_test(lambda u: u.is_superuser)
def superadmin_import_employees(request, company_id):
    if request.method == 'POST':
        company = get_object_or_404(Company, id=company_id)
        employee_file = request.FILES.get('employee_list')
        
        if employee_file:
            from django.db import transaction
            import csv
            import io
            
            try:
                with transaction.atomic():
                    file_ext = employee_file.name.split('.')[-1].lower()
                    added_count = 0
                    
                    if file_ext == 'csv':
                        file_data = employee_file.read().decode('utf-8-sig')
                        sniffer = csv.Sniffer()
                        try:
                            dialect = sniffer.sniff(file_data[:1024])
                            csv_data = csv.reader(io.StringIO(file_data), dialect)
                        except csv.Error:
                            if ';' in file_data[:1024]:
                                csv_data = csv.reader(io.StringIO(file_data), delimiter=';')
                            else:
                                csv_data = csv.reader(io.StringIO(file_data), delimiter=',')
                                
                        next(csv_data, None) # Skip header
                        for row in csv_data:
                            if len(row) >= 5:
                                first_name = row[0].strip()
                                last_name = row[1].strip()
                                email = row[2].strip()
                                password = row[3].strip()
                                team_name = row[4].strip()
                                
                                if not email: continue
                                
                                user, created = User.objects.get_or_create(email=email, defaults={
                                    'username': email,
                                    'first_name': first_name,
                                    'last_name': last_name,
                                })
                                if created or not user.has_usable_password():
                                    user.set_password(password)
                                    user.save()
                                    
                                department = None
                                if team_name:
                                    department, _ = Department.objects.get_or_create(name=team_name, company=company)
                                    
                                profile, _ = Profile.objects.get_or_create(user=user)
                                profile.company_link = company
                                if department:
                                    profile.department = department
                                profile.save()
                                added_count += 1
                                
                    elif file_ext in ['xlsx', 'xls']:
                        try:
                            import openpyxl
                            wb = openpyxl.load_workbook(employee_file)
                            sheet = wb.active
                            for row in sheet.iter_rows(min_row=2, values_only=True):
                                if row and len(row) >= 5 and row[2]:
                                    first_name = str(row[0]).strip() if row[0] else ''
                                    last_name = str(row[1]).strip() if row[1] else ''
                                    email = str(row[2]).strip()
                                    password = str(row[3]).strip() if row[3] else ''
                                    team_name = str(row[4]).strip() if row[4] else ''
                                    
                                    user, created = User.objects.get_or_create(email=email, defaults={
                                        'username': email,
                                        'first_name': first_name,
                                        'last_name': last_name,
                                    })
                                    if created or not user.has_usable_password():
                                        user.set_password(password)
                                        user.save()
                                        
                                    department = None
                                    if team_name:
                                        department, _ = Department.objects.get_or_create(name=team_name, company=company)
                                        
                                    profile, _ = Profile.objects.get_or_create(user=user)
                                    profile.company_link = company
                                    if department:
                                        profile.department = department
                                    profile.save()
                                    added_count += 1
                        except ImportError:
                            messages.warning(request, 'Nepavyko apdoroti Excel failo. Instaliuokite "openpyxl".')
                            return redirect('superadmin_edit_employees', company_id=company_id)
                            
                    messages.success(request, f'Sėkmingai importuota / atnaujinta {added_count} darbuotojų iš sąrašo.')
            except Exception as e:
                messages.error(request, f'Klaida apdorojant failą: {str(e)}')
        else:
            messages.error(request, 'Nepasirinktas joks failas.')
            
    return redirect('superadmin_edit_employees', company_id=company_id)


@user_passes_test(lambda u: u.is_superuser)
def superadmin_delete_department(request, company_id, department_id):
    if request.method == 'POST':
        company = get_object_or_404(Company, id=company_id)
        department = get_object_or_404(Department, id=department_id, company=company)
        
        department.delete()
        messages.success(request, f'Departamentas "{department.name}" sėkmingai ištrintas.')
            
    return redirect('superadmin_edit_hierarchy', company_id=company_id)

@user_passes_test(lambda u: u.is_superuser)
def superadmin_edit_department(request, company_id, department_id):
    company = get_object_or_404(Company, id=company_id)
    department = get_object_or_404(Department, id=department_id, company=company)
    
    if request.method == 'POST':
        form = DepartmentForm(request.user, request.POST, instance=department)
        # Override parent and manager querysets to target company's departments, excluding self to prevent loop
        form.fields['parent'].queryset = Department.objects.filter(company=company).exclude(id=department.id)
        form.fields['manager'].queryset = User.objects.filter(profile__company_link=company)
        
        if form.is_valid():
            dept = form.save()
            if dept.manager:
                manager_profile = dept.manager.profile
                manager_profile.department = dept
                manager_profile.save()
            messages.success(request, f'Departamentas "{dept.name}" sėkmingai atnaujintas.')
            return redirect('superadmin_edit_hierarchy', company_id=company_id)
    else:
        form = DepartmentForm(request.user, instance=department)
        form.fields['parent'].queryset = Department.objects.filter(company=company).exclude(id=department.id)
        form.fields['manager'].queryset = User.objects.filter(profile__company_link=company)

    context = {
        'company': company,
        'department': department,
        'form': form,
    }
    return render(request, 'superadmin/edit_department.html', context)

@user_passes_test(lambda u: u.is_superuser)
def superadmin_toggle_admin(request, company_id, user_id):
    if request.method == 'POST':
        company = get_object_or_404(Company, id=company_id)
        target_user = get_object_or_404(User, id=user_id)
        if hasattr(target_user, 'profile') and target_user.profile.company_link == company:
            target_user.profile.is_company_admin = not target_user.profile.is_company_admin
            target_user.profile.save()
            status = 'priskirtas' if target_user.profile.is_company_admin else 'pašalintas iš'
            messages.success(request, f'{target_user.get_full_name()} {status} administratorių.')
    return redirect('superadmin_company_detail', company_id=company_id)



@user_passes_test(lambda u: u.is_superuser)
def superadmin_billing_overview(request):
    """
    Bendra sąskaitų statistika visoms įmonėms.
    Kiekvienai įmonei skaičiuoja dabartinio mėnesio sąskaitą (Max Count).
    """
    from users.models import ContractSettings
    from users.billing_service import calculate_monthly_bill
    import calendar

    today = date.today()
    try:
        selected_year = int(request.GET.get('year', today.year))
        selected_month = int(request.GET.get('month', today.month))
        if not (1 <= selected_month <= 12):
            selected_month = today.month
    except (ValueError, TypeError):
        selected_year, selected_month = today.year, today.month

    companies = Company.objects.all().order_by('name')

    rows = []
    total_amount = 0
    total_employees = 0
    companies_with_settings = 0
    companies_without_settings = 0

    for company in companies:
        bill = calculate_monthly_bill(company.id, selected_year, selected_month)
        if bill.get('has_settings'):
            companies_with_settings += 1
            total_amount += bill['final_amount']
            total_employees += bill['max_count']
        else:
            companies_without_settings += 1
        rows.append({
            'company': company,
            'bill': bill,
        })

    # ── Grafiko duomenys: per įmonę, pasirinktas laikotarpis ──────────────────
    import json

    # Laikotarpis: praėję + ateities mėnesiai
    chart_past = int(request.GET.get('chart_past', 6))   # praėjusių mėnesių
    chart_future = int(request.GET.get('chart_future', 3))  # prognozuojamų
    chart_past = max(1, min(chart_past, 24))
    chart_future = max(0, min(chart_future, 12))

    # Sugeneruojame visus mėnesius kairė→dešinė: seniausias...dabartinis...ateitis
    chart_labels = []
    chart_is_future = []

    # Pradžios taškas: chart_past mėnesių atgal nuo šiandien
    start_m = today.month - chart_past
    start_y = today.year
    while start_m <= 0:
        start_m += 12
        start_y -= 1

    m, y = start_m, start_y
    total_months = chart_past + 1 + chart_future
    for _ in range(total_months):
        label = f"{y}-{m:02d}"
        chart_labels.append({'label': label, 'year': y, 'month': m})
        chart_is_future.append(date(y, m, 1) > date(today.year, today.month, 1))
        # Pereiti į kitą mėnesį
        m += 1
        if m > 12:
            m = 1
            y += 1

    # Vieno mėnesio prognozė = paskutinio žinomo mėnesio sąskaitos sumos
    # Sukaupiam duomenis pagal įmonę
    from users.billing_service import calculate_monthly_bill as _calc

    # Spalvų paletė įmonėms
    PALETTE = [
        '#2d4a77', '#3b7dd8', '#52a8ff', '#7ec8e3',
        '#a78bfa', '#34d399', '#f59e0b', '#f87171',
        '#94a3b8', '#fb923c', '#e879f9', '#4ade80',
    ]

    companies_with_contract = [c for c in companies if hasattr(c, 'contract_settings')]
    companies_list = list(companies)

    chart_datasets = []
    for idx, company in enumerate(companies_list):
        color = PALETTE[idx % len(PALETTE)]
        values = []
        for slot in chart_labels:
            b = _calc(company.id, slot['year'], slot['month'])
            if b.get('has_settings'):
                values.append(float(b['final_amount']))
            else:
                values.append(0.0)
        chart_datasets.append({
            'label': company.name,
            'data': values,
            'backgroundColor': color + 'cc',   # slight transparency
            'borderColor': color,
            'borderWidth': 1,
            'borderRadius': 4,
        })

    label_strs = [s['label'] for s in chart_labels]

    # Mėnesių sąrašas pasirinkimui (paskutiniai 12 mėnesių, naujesni pirmiau)
    month_options = []
    for i in range(12):
        m = today.month - i
        y = today.year
        if m <= 0:
            m += 12
            y -= 1
        month_options.append({'year': y, 'month': m, 'label': f"{y} {calendar.month_abbr[m]}"})

    context = {
        'rows': rows,
        'selected_year': selected_year,
        'selected_month': selected_month,
        'selected_month_label': f"{selected_year} {calendar.month_name[selected_month]}",
        'month_options': month_options,
        'total_amount': total_amount,
        'total_employees': total_employees,
        'companies_with_settings': companies_with_settings,
        'companies_without_settings': companies_without_settings,
        'today': today,
        'chart_labels_json': json.dumps(label_strs),
        'chart_datasets_json': json.dumps(chart_datasets),
        'chart_future_json': json.dumps(chart_is_future),
        'chart_past': chart_past,
        'chart_future': chart_future,
        # Mygtukų variantai (value, label)
        'past_options': [(3, '3M'), (6, '6M'), (12, '12M'), (24, '24M')],
        'future_options': [(0, 'Išj.'), (1, '+1M'), (3, '+3M'), (6, '+6M')],
    }
    return render(request, 'superadmin/billing_overview.html', context)



@user_passes_test(lambda u: u.is_superuser)
def superadmin_company_billing(request, company_id):
    """
    Rodo įmonės sutarčių sąrašą, leidžia kurti naujas sutartis
    ir skaičiuoja mėnesio sąskaitą pagal Max Count strategiją.
    """
    from users.models import ContractSettings, EmployeeCountLog
    from users.billing_service import calculate_monthly_bill
    from decimal import Decimal, InvalidOperation
    import calendar

    company = get_object_or_404(Company, id=company_id)

    # ── POST: sukurti naują sutartį ───────────────────────────────────────────
    if request.method == 'POST':
        action = request.POST.get('action', 'create')

        if action == 'delete':
            contract_id = request.POST.get('contract_id')
            ContractSettings.objects.filter(id=contract_id, company=company).delete()
            messages.success(request, 'Sutartis ištrinta.')
            return redirect('superadmin_company_billing', company_id=company_id)

        # Nauja sutartis
        price_str = request.POST.get('price_per_employee', '').strip()
        min_fee_str = request.POST.get('minimum_fee', '0').strip() or '0'
        contract_start = request.POST.get('contract_start', '').strip()
        contract_end = request.POST.get('contract_end', '').strip() or None

        try:
            price = Decimal(price_str)
            min_fee = Decimal(min_fee_str)
        except InvalidOperation:
            messages.error(request, 'Neteisingas kainos formatas. Naudokite skaičius, pvz.: 9.99')
            return redirect('superadmin_company_billing', company_id=company_id)

        if not contract_start:
            messages.error(request, 'Sutarties pradžios data yra privaloma.')
            return redirect('superadmin_company_billing', company_id=company_id)

        ContractSettings.objects.create(
            company=company,
            price_per_employee=price,
            minimum_fee=min_fee,
            contract_start=contract_start,
            contract_end=contract_end,
        )
        messages.success(request, 'Nauja sutartis sėkmingai sukurta.')
        return redirect('superadmin_company_billing', company_id=company_id)

    # ── GET ───────────────────────────────────────────────────────────────────
    today = date.today()
    try:
        selected_year = int(request.GET.get('year', today.year))
        selected_month = int(request.GET.get('month', today.month))
        if not (1 <= selected_month <= 12):
            selected_month = today.month
    except (ValueError, TypeError):
        selected_year, selected_month = today.year, today.month

    # Visos šios įmonės sutartys (naujiausios pirmiau)
    all_contracts = ContractSettings.objects.filter(company=company).order_by('-contract_start')

    # Skaičiuojame sąskaitą pasirinktam mėnesiui
    bill = calculate_monthly_bill(company_id, selected_year, selected_month)

    # Paskutiniai EmployeeCountLog įrašai
    recent_logs = EmployeeCountLog.objects.filter(company=company).order_by('-recorded_at')[:15]

    # Mėnesių sąrašas: nuo anksčiausios sutarties pradžios iki dabar
    month_options = []
    earliest = all_contracts.order_by('contract_start').first()
    if earliest:
        iter_date = date(earliest.contract_start.year, earliest.contract_start.month, 1)
        current = date(today.year, today.month, 1)
        while iter_date <= current:
            month_options.append({
                'year': iter_date.year,
                'month': iter_date.month,
                'label': f"{iter_date.year}-{iter_date.month:02d}",
            })
            if iter_date.month == 12:
                iter_date = date(iter_date.year + 1, 1, 1)
            else:
                iter_date = date(iter_date.year, iter_date.month + 1, 1)
        month_options.reverse()

    context = {
        'company': company,
        'bill': bill,
        'all_contracts': all_contracts,
        'recent_logs': recent_logs,
        'selected_year': selected_year,
        'selected_month': selected_month,
        'selected_month_label': f"{selected_year}-{selected_month:02d}",
        'month_options': month_options,
        'today': today,
    }
    return render(request, 'superadmin/company_billing.html', context)

@user_passes_test(lambda u: u.is_superuser)
def superadmin_remove_employee(request, company_id, user_id):
    if request.method == 'POST':
        user = get_object_or_404(User, id=user_id)
        if hasattr(user, 'profile') and user.profile.company_link_id == company_id:
            user.profile.company_link = None
            user.profile.department = None
            user.profile.manager = None
            user.profile.save()
            messages.success(request, f'Vartotojas {user.email} pašalintas iš įmonės.')
        else:
            messages.error(request, 'Vartotojas nepriklauso šiai įmonei.')
            
    return redirect('superadmin_edit_employees', company_id=company_id)

@user_passes_test(lambda u: u.is_superuser)
def superadmin_impersonate_user(request, user_id):
    original_user_id = request.user.id
    target_user = get_object_or_404(User, id=user_id)
    
    # Capture company ID to return to it later
    company_id = None
    if hasattr(target_user, 'profile') and target_user.profile.company_link:
        company_id = target_user.profile.company_link.id

    # Log in as the target user without authentication backend check
    # We specify the backend manually to bypass authentication
    login(request, target_user, backend='django.contrib.auth.backends.ModelBackend')

    # Save the original user's ID in the session AFTER login because login flushes session
    request.session['impersonator_id'] = original_user_id
    if company_id:
        request.session['impersonation_return_company_id'] = company_id
    
    return redirect('home')

def stop_impersonation(request):
    impersonator_id = request.session.get('impersonator_id')
    return_company_id = request.session.get('impersonation_return_company_id')
    
    if impersonator_id:
        original_user = get_object_or_404(User, id=impersonator_id)
        
        # Log in back as the original user
        login(request, original_user, backend='django.contrib.auth.backends.ModelBackend')
        
        # Remove the impersonator data from the session safely
        request.session.pop('impersonator_id', None)
        request.session.pop('impersonation_return_company_id', None)
        
        if return_company_id:
            return redirect('superadmin_edit_employees', company_id=return_company_id)
        return redirect('superadmin_dashboard')
        
    return redirect('home')

# === Questionnaires ===

@login_required
def questionnaires_list(request):
    from .models import Questionnaire, Trait
    from users.models import Department
    
    if not Trait.objects.exists():
        default_traits = [
            'Komandinis darbas', 'Komunikabilumas', 'Iniciatyvumas', 'Problemų sprendimas', 'Lyderystė',
            'Analitinis mąstymas', 'Kūrybiškumas', 'Adaptabilumas', 'Atsakingumas', 'Laiko planavimas',
            'Techninės žinios', 'Strateginis mąstymas', 'Klientų aptarnavimas', 'Derybos', 'Prezentavimo įgūdžiai',
            'Patikimumas', 'Motyvacija', 'Pozityvumas', 'Efektyvumas', 'Savarankiškumas'
        ]
        for t in default_traits:
            Trait.objects.get_or_create(name=t)
            
    questionnaires = Questionnaire.objects.filter(created_by=request.user).order_by('-created_at')
    all_traits = Trait.objects.all().order_by('name')
    
    # Get team members for sending questionnaires
    team_members_qs = User.objects.none()
    try:
        user_company_link = request.user.profile.company_link
        if user_company_link:
            team_members_qs = User.objects.filter(profile__company_link=user_company_link).exclude(id=request.user.id)
        else:
             team_members_qs = User.objects.filter(profile__company_link__isnull=True).exclude(id=request.user.id)
    except Exception:
        team_members_qs = User.objects.filter(profile__isnull=True).exclude(id=request.user.id)
        
    team_members = team_members_qs.order_by('first_name', 'last_name')

    managed_departments = list(Department.objects.filter(manager=request.user).order_by('name'))
    sub_departments = list(Department.objects.filter(parent__in=managed_departments).order_by('name'))
    all_managed = list({d.id: d for d in managed_departments + sub_departments}.values())
    all_managed.sort(key=lambda x: x.name)

    return render(request, 'questionnaires/list.html', {
        'questionnaires': questionnaires,
        'all_traits': all_traits,
        'team_members': team_members,
        'managed_departments': all_managed,
        'is_company_active': is_company_active(request.user)
    })

@login_required
def create_questionnaire(request):
    from .models import Questionnaire, Trait
    if not is_company_active(request.user):
        messages.error(request, 'Jūsų įmonė yra išjungta. Veiksmas negalimas.')
        return redirect('questionnaires_list')
        
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        if not title:
            messages.error(request, 'Klausimyno pavadinimas yra privalomas.')
            return redirect('questionnaires_list')

        questionnaire = Questionnaire.objects.create(title=title, created_by=request.user)

        # Add existing traits
        trait_ids = request.POST.getlist('trait_ids')
        for tid in trait_ids:
            try:
                trait = Trait.objects.get(id=int(tid))
                questionnaire.traits.add(trait)
            except (Trait.DoesNotExist, ValueError):
                pass

        # Add custom traits
        custom_traits = request.POST.getlist('custom_traits')
        for name in custom_traits:
            name = name.strip()
            if name:
                trait, _ = Trait.objects.get_or_create(name=name, defaults={'created_by': request.user})
                questionnaire.traits.add(trait)

        messages.success(request, f'Klausimynas "{title}" sėkmingai sukurtas!')
    return redirect('questionnaires_list')


@login_required
def create_team_questionnaire(request):
    from .models import Questionnaire, Trait, FeedbackRequest
    from users.models import Department
    if not is_company_active(request.user):
        messages.error(request, 'Jūsų įmonė yra išjungta. Veiksmas negalimas.')
        return redirect('questionnaires_list')
        
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        department_id = request.POST.get('department_id')
        
        if not title or not department_id:
            messages.error(request, 'Pavadinimas ir komanda yra privalomi.')
            return redirect('questionnaires_list')

        try:
            department = Department.objects.get(id=int(department_id))
            # Just ensuring user can manage it or its parent
            managed_ids = list(Department.objects.filter(manager=request.user).values_list('id', flat=True))
            sub_ids = list(Department.objects.filter(parent_id__in=managed_ids).values_list('id', flat=True))
            if department.id not in managed_ids and department.id not in sub_ids:
                raise ValueError
        except (Department.DoesNotExist, ValueError):
            messages.error(request, 'Nepavyko rasti pasirinktos komandos arba neturite jai teisių.')
            return redirect('questionnaires_list')

        questionnaire = Questionnaire.objects.create(title=title, created_by=request.user, is_team=True, target_department=department)

        # Add existing traits
        trait_ids = request.POST.getlist('trait_ids')
        for tid in trait_ids:
            try:
                trait = Trait.objects.get(id=int(tid))
                questionnaire.traits.add(trait)
            except (Trait.DoesNotExist, ValueError):
                pass

        # Add custom traits
        custom_traits = request.POST.getlist('custom_traits')
        for name in custom_traits:
            name = name.strip()
            if name:
                trait, _ = Trait.objects.get_or_create(name=name, defaults={'created_by': request.user})
                questionnaire.traits.add(trait)

        messages.success(request, f'Komandinė forma "{title}" sėkmingai sukurta!')
    return redirect('questionnaires_list')


@login_required
@require_POST
def send_questionnaire(request):
    from .models import Questionnaire, FeedbackRequest
    from django.contrib.auth.models import User
    from datetime import date, timedelta
    
    if not is_company_active(request.user):
        messages.error(request, 'Jūsų įmonė yra išjungta. Veiksmas negalimas.')
        return redirect('questionnaires_list')
        
    questionnaire_id = request.POST.get('questionnaire_id')
    colleague_ids = request.POST.getlist('colleague_ids')
    
    if not questionnaire_id or not colleague_ids:
        messages.error(request, 'Trūksta duomenų klausimyno siuntimui. Įsitikinkite, kad pasirinkote bent vieną kolegą.')
        return redirect('questionnaires_list')
        
    questionnaire = get_object_or_404(Questionnaire, id=questionnaire_id, created_by=request.user)
    
    project_name = questionnaire.title
    if questionnaire.is_team and getattr(questionnaire, 'target_department', None):
        project_name = f"{questionnaire.title} ({questionnaire.target_department.name})"

    sent_count = 0
    skipped_count = 0

    for c_id in colleague_ids:
        requested_to = get_object_or_404(User, id=c_id)
        
        if requested_to.profile.company_link != request.user.profile.company_link:
            skipped_count += 1
            continue
        
        # Check if a pending request already exists for this questionnaire and this colleague
        existing = FeedbackRequest.objects.filter(
            requester=request.user,
            requested_to=requested_to,
            status='pending',
            project_name=project_name
        ).exists()
        
        if existing:
            skipped_count += 1
            continue
            
        FeedbackRequest.objects.create(
            requester=request.user,
            requested_to=requested_to,
            project_name=project_name,
            questionnaire=questionnaire,
            comment=f"Prašau užpildyti klausimyną: {questionnaire.title}",
            due_date=date.today() + timedelta(days=7)
        )
        sent_count += 1
        
    if sent_count > 0 and skipped_count == 0:
        messages.success(request, f'Klausimynas "{questionnaire.title}" sėkmingai išsiųstas pasirinktiems ({sent_count}) kolegoms!')
    elif sent_count > 0 and skipped_count > 0:
        messages.success(request, f'Klausimynas išsiųstas {sent_count} kolegom(s). Praleisti {skipped_count}, nes jie jau turi aktyvią šios formos užklausą.')
    elif sent_count == 0 and skipped_count > 0:
        messages.warning(request, f'Klausimynas nebuvo išsiųstas, nes visi pasirinkti kolegos jau turi aktyvią šios formos užklausą.')
        
    return redirect('questionnaires_list')


@login_required
def edit_questionnaire(request, questionnaire_id):
    from .models import Questionnaire, Trait # Added import here for edit_questionnaire
    questionnaire = get_object_or_404(Questionnaire, id=questionnaire_id, created_by=request.user)
    
    if request.method == 'POST':
        title = request.POST.get('title')
        trait_ids = request.POST.getlist('trait_ids')
        custom_traits = request.POST.getlist('custom_traits')
        
        if not title:
            messages.error(request, 'Klausimyno pavadinimas yra privalomas.')
            return redirect('questionnaires_list')
            
        questionnaire.title = title
        questionnaire.save()
        
        # Clear existing traits
        questionnaire.traits.clear()
        
        # Add selected existing traits
        for trait_id in trait_ids:
            try:
                trait = Trait.objects.get(id=trait_id)
                questionnaire.traits.add(trait)
            except Trait.DoesNotExist:
                continue
                
        # Add new custom traits
        for trait_name in custom_traits:
            name = trait_name.strip()
            if name:
                trait, created = Trait.objects.get_or_create(
                    name=name,
                    defaults={'created_by': request.user}
                )
                questionnaire.traits.add(trait)
                
        messages.success(request, 'Klausimynas sėkmingai atnaujintas.')
        return redirect('questionnaires_list')
        
    messages.error(request, 'Klaida atnaujinant klausimyną.')
    return redirect('questionnaires_list')

@login_required
def delete_questionnaire(request, questionnaire_id):
    from .models import Questionnaire
    if request.method == 'POST':
        questionnaire = get_object_or_404(Questionnaire, id=questionnaire_id, created_by=request.user)
        questionnaire.delete()
        messages.success(request, 'Klausimynas sėkmingai ištrintas.')
    return redirect('questionnaires_list')

@login_required
def questionnaire_statistics(request, questionnaire_id):
    from .models import Questionnaire, FeedbackRequest, Feedback
    from .services import FeedbackAnalytics

    questionnaire = get_object_or_404(Questionnaire, id=questionnaire_id, created_by=request.user)

    # Fetch feedback requests related to this questionnaire
    # For now, we linked them by using project_name=questionnaire.title
    feedback_requests = FeedbackRequest.objects.filter(
        requester=request.user,
        project_name=questionnaire.title,
        status='completed'
    ).select_related('requested_to')
    
    feedbacks = Feedback.objects.filter(feedback_request__in=feedback_requests)
    
    overall_avg_rating = feedbacks.aggregate(Avg('rating'))['rating__avg'] or 0
    received_feedback_count = feedbacks.count()
    
    from .models import TraitRating
    from collections import defaultdict

    traits = questionnaire.traits.all()
    trait_ratings = TraitRating.objects.filter(feedback__in=feedbacks)
    
    competencies = []
    
    # Pre-fetch trait ratings per date for chart
    trait_ratings_by_date = defaultdict(lambda: defaultdict(list))
    for tr in trait_ratings.select_related('feedback__feedback_request'):
        d = tr.feedback.feedback_request.created_at.date()
        trait_ratings_by_date[tr.trait_id][d].append(tr.rating)

    for trait in traits:
        # Average score for this trait
        avg = trait_ratings.filter(trait=trait).aggregate(Avg('rating'))['rating__avg'] or 0
        competencies.append({
            'name': trait.name,
            'score': round(avg, 2)
        })

    all_keywords = []
    for fb in feedbacks:
        if fb.keywords:
            keys = [k.strip() for k in fb.keywords.split(',') if k.strip()]
            all_keywords.extend(keys)

    all_keywords = list(set(all_keywords)) # Unique keywords

    strengths = []
    improvements = []
    
    # We can separate comments into strengths/improvements if we want, but for now we'll just list comments
    for fb in feedbacks:
        if fb.comments:
            strengths.append(fb.comments) 

    import json
    from django.db.models.functions import TruncDate

    chronological_feedbacks = feedbacks.annotate(
        date=TruncDate('feedback_request__created_at')
    ).values('date').annotate(
        avg_rating=Avg('rating')
    ).order_by('date')

    chart_labels = [fb['date'].strftime('%Y-%m-%d') if fb['date'] else 'Data nežinoma' for fb in chronological_feedbacks]
    
    chart_datasets = [
        {'label': 'Bendras', 'data': [round(fb['avg_rating'], 2) for fb in chronological_feedbacks], 'borderColor': '#8B5CF6', 'tension': 0.3},
    ]
    
    colors = ['#3B82F6', '#10B981', '#F59E0B', '#EF4444', '#6366F1', '#EC4899', '#14B8A6', '#F43F5E']
    
    for i, trait in enumerate(traits):
        data = []
        for fb_dict in chronological_feedbacks:
            d = fb_dict['date']
            # if d is None, fallback
            ratings = trait_ratings_by_date[trait.id].get(d, [])
            avg = sum(ratings) / len(ratings) if ratings else 0.0
            data.append(round(avg, 2))
            
        chart_datasets.append({
            'label': trait.name,
            'data': data,
            'borderColor': colors[i % len(colors)],
            'tension': 0.3,
            'hidden': True
        })

    chart_data = {
        'labels': chart_labels,
        'datasets': chart_datasets
    }

    context = {
        'questionnaire': questionnaire,
        'overall_avg_rating': round(overall_avg_rating, 2),
        'received_feedback_count': received_feedback_count,
        'competencies': competencies,
        'all_keywords': all_keywords,
        'strengths': strengths,
        'improvements': improvements,
        'chart_data_json': json.dumps(chart_data),
    }

    
    return render(request, 'questionnaires/statistics.html', context)


# ==========================================
# SUPERUSERS MANAGEMENT (SUPERADMIN)
# ==========================================

@user_passes_test(lambda u: u.is_superuser)
def superadmin_superusers_list(request):
    superusers = User.objects.filter(is_superuser=True).order_by('-date_joined')
    return render(request, 'superadmin/superusers_list.html', {'superusers': superusers})

@user_passes_test(lambda u: u.is_superuser)
def superadmin_create_superuser(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        password = request.POST.get('password')

        if not email or not password:
            messages.error(request, 'El. paštas ir slaptažodis yra privalomi.')
            return render(request, 'superadmin/superuser_form.html')

        if User.objects.filter(username=email).exists() or User.objects.filter(email=email).exists():
            messages.error(request, 'Vartotojas su tokiu el. paštu jau egzistuoja.')
            return render(request, 'superadmin/superuser_form.html')

        user = User.objects.create_superuser(
            username=email,
            email=email,
            password=password,
            first_name=first_name,
            last_name=last_name
        )
        messages.success(request, f'Supervartotojas {email} sėkmingai sukurtas.')
        return redirect('superadmin_superusers_list')

    return render(request, 'superadmin/superuser_form.html')

@user_passes_test(lambda u: u.is_superuser)
def superadmin_edit_superuser(request, user_id):
    superuser = get_object_or_404(User, id=user_id, is_superuser=True)
    
    if request.method == 'POST':
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        password = request.POST.get('password')

        superuser.first_name = first_name
        superuser.last_name = last_name
        
        if password:
            superuser.set_password(password)
            
        superuser.save()
        messages.success(request, f'Supervartotojo {superuser.email} duomenys atnaujinti.')
        return redirect('superadmin_superusers_list')

    return render(request, 'superadmin/superuser_form.html', {'superuser': superuser})

@user_passes_test(lambda u: u.is_superuser)
def superadmin_delete_superuser(request, user_id):
    if request.method == 'POST':
        superuser = get_object_or_404(User, id=user_id, is_superuser=True)
        if superuser == request.user:
            messages.error(request, 'Negalite ištrinti patys savęs.')
        else:
            superuser.delete()
            messages.success(request, 'Supervartotojas sėkmingai ištrintas.')
    return redirect('superadmin_superusers_list')

@user_passes_test(lambda u: u.is_superuser)
def superadmin_users_list(request):
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        company_id = request.POST.get('company_id')
        
        if user_id and company_id:
            target_user = get_object_or_404(User, id=user_id, is_superuser=False)
            
            if company_id == 'none':
                target_user.profile.company_link = None
            else:
                company = get_object_or_404(Company, id=company_id)
                target_user.profile.company_link = company
                
            target_user.profile.save()
            messages.success(request, f'Vartotojo {target_user.email} įmonė atnaujinta.')
            
        return redirect('superadmin_users_list')

    sys_users = User.objects.filter(is_superuser=False).select_related('profile', 'profile__company_link').order_by('-date_joined')
    companies = Company.objects.all().order_by('name')
    
    context = {
        'users': sys_users,
        'companies': companies
    }
    
    return render(request, 'superadmin/users_list.html', context)


from django.contrib.admin.views.decorators import staff_member_required
from .models import GlobalSettings

@staff_member_required
def superadmin_features(request):
    if not request.user.is_superuser:
        return redirect('home')
        
    settings = GlobalSettings.load()
    
    if request.method == 'POST':
        settings.personal_form_enabled = request.POST.get('personal_form_enabled') == 'on'
        settings.team_form_enabled = request.POST.get('team_form_enabled') == 'on'
        settings.save()
        messages.success(request, 'Funkcionalumų nustatymai sėkmingai atnaujinti.')
        return redirect('superadmin_features')
        
    return render(request, 'superadmin/features.html', {'settings': settings})
